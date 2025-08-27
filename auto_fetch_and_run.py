#!/usr/bin/env python3
"""
Auto-fetch ABS Retail Trade (Tables 1 & 12) and ASX200 EOM values, then compute the AU Lipstick Index.

Run:
  python auto_fetch_and_run.py

Optional .env (place next to this file):
  LIP_BASE_DIR=.
  # or explicitly:
  # LIP_BASE_DIR=/absolute/path/where/output/should/live
  # LIP_DATA_DIR=/absolute/path/override/for/data
  # LIP_OUT_DIR=/absolute/path/override/for/output

Requires (install if needed):
  pip install requests beautifulsoup4 lxml pandas openpyxl matplotlib python-dotenv

What it does:
  1) Pulls ABS "Retail Trade, Australia — latest release" and finds:
       - Table 1 (by industry group)  -> Clothing, footwear & personal accessory retailing (SA, $m)
       - Table 12 (state by industry subgroup, SA) -> Pharmaceutical, cosmetic & toiletry goods retailing
     Robust to page wording; prefers ABS-pack filenames (850101.xlsx, 8501012.xlsx) and falls back to inspecting workbook text.

  2) Gets ASX "S&P/ASX 200" end-of-month (EOM close):
       - Try official ASX page
       - Fallback to Yahoo Finance (^AXJO) CSV + HTML history (with retries/backoff)
       - Fallback to Stooq monthly CSV
       - Fallback to MarketWatch (XJO) CSV (daily → resample to month-end)
       - Fallback to Investing.com monthly HTML table
       - If all fail, create/use local cache data/asx200_eom.csv (template generated if missing)

  3) Builds LipstickIndex_t = 0.60 * MoM(Cosmetics) + 0.40 * MoM(Clothing), flags months where:
       LipstickIndex_t > 0  AND  ASX200_MoM < 0

Outputs:
  data/clothing_sa.csv
  data/cosmetics_sa.csv
  data/asx200_eom.csv
  data/lipstick_divergences.csv
  output/lipstick_vs_asx_timeseries.png
  output/lipstick_scatter.png
  data/table1.xlsx, data/table12.xlsx (downloaded source workbooks)
"""

import os, re, io, sys, datetime as dt, tempfile, random, csv
import pandas as pd
import numpy as np
import requests
from bs4 import BeautifulSoup
import matplotlib.pyplot as plt

# ----------------------------
# .env support (optional)
# ----------------------------
def _load_env():
    try:
        from dotenv import load_dotenv
        load_dotenv()
    except Exception:
        # python-dotenv not installed; ignore quietly
        pass

_load_env()

def _env_path(name: str, default: str) -> str:
    val = os.environ.get(name, "").strip()
    if val:
        return os.path.abspath(os.path.expanduser(val))
    return default

# ----------------------------
# Config / Paths (env-overridable)
# ----------------------------
BASE = _env_path("LIP_BASE_DIR", os.path.dirname(__file__))
DATA = _env_path("LIP_DATA_DIR", os.path.join(BASE, "data"))
OUT  = _env_path("LIP_OUT_DIR",  os.path.join(BASE, "output"))

def _ensure_dirs():
    for d in [BASE, DATA, OUT]:
        try:
            os.makedirs(d, exist_ok=True)
        except Exception as e:
            print(f"WARNING: Could not create directory {d} ({e}). Falling back to temp.")
            tmp = os.path.join(tempfile.gettempdir(), "lipstick_index_fallback")
            os.makedirs(tmp, exist_ok=True)
            return (tmp, tmp, tmp)
    return (BASE, DATA, OUT)

BASE, DATA, OUT = _ensure_dirs()

ABS_LATEST_URL = "https://www.abs.gov.au/statistics/industry/retail-and-wholesale-trade/retail-trade-australia/latest-release"
ASX_HIST_URL   = "https://www.asx.com.au/about/market-statistics/historical-market-statistics"
YF_CSV_HOSTS   = ["query1.finance.yahoo.com", "query2.finance.yahoo.com"]
YF_HTML_URL    = "https://finance.yahoo.com/quote/%5EAXJO/history?frequency=1mo&filter=history"
MW_CSV_BASE    = "https://www.marketwatch.com/investing/index/xjo/downloaddatapartial"
INV_HTML_URL   = "https://www.investing.com/indices/aus-200-historical-data"

# Weights for composite
WEIGHT_COSMETICS = 0.60
WEIGHT_CLOTHING  = 0.40

# ----------------------------
# Helpers (HTTP / parsing / safe save)
# ----------------------------
def _ua():
    return random.choice([
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17 Safari/605.1.15",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119 Safari/537.36",
    ])

def _get(url, **kwargs):
    headers = kwargs.pop("headers", {})
    headers.setdefault("User-Agent", _ua())
    return requests.get(url, headers=headers, timeout=kwargs.pop("timeout", 30), **kwargs)

def _get_soup(url):
    r = _get(url)
    r.raise_for_status()
    try:
        return BeautifulSoup(r.text, "lxml")
    except Exception:
        return BeautifulSoup(r.text, "html.parser")

def _abs_make_absolute(href: str) -> str:
    if href.startswith("http"):
        return href
    return "https://www.abs.gov.au" + href

def _xlsx_text_fingerprint(xbytes: bytes) -> str:
    """Return a lowercase blob of text from the first ~50 rows of each sheet to score candidates."""
    try:
        xl = pd.ExcelFile(io.BytesIO(xbytes))
        texts = []
        for sheet in xl.sheet_names:
            try:
                df = xl.parse(sheet, header=None, nrows=50)
                texts.extend(df.astype(str).fillna("").values.ravel().tolist())
            except Exception:
                continue
        return " ".join(texts).lower()
    except Exception:
        return ""

def eom_date_from_text(month_name: str, year: str) -> pd.Timestamp:
    """Map e.g. 'Oct' + '2024' -> 2024-10-31 (end-of-month)"""
    month_num = dt.datetime.strptime(month_name[:3], "%b").month
    if month_num == 12:
        last = dt.date(int(year), 12, 31)
    else:
        last = dt.date(int(year), month_num + 1, 1) - dt.timedelta(days=1)
    return pd.Timestamp(last)

def _safe_to_csv(df: pd.DataFrame, path: str, max_attempts: int = 2):
    """
    Try writing to path. If PermissionError (e.g., locked by Excel/OneDrive),
    fallback to a temp folder and print where it went.
    """
    attempt = 0
    while attempt < max_attempts:
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            df.to_csv(path, index=False)
            return path
        except PermissionError:
            attempt += 1
    # Fallback
    alt_dir = os.path.join(tempfile.gettempdir(), "lipstick_index_fallback")
    os.makedirs(alt_dir, exist_ok=True)
    alt_path = os.path.join(alt_dir, os.path.basename(path))
    df.to_csv(alt_path, index=False)
    print(f"WARNING: Could not write to {path} (possibly locked). Saved instead to: {alt_path}")
    return alt_path

def _safe_write_bytes(b: bytes, path: str, max_attempts: int = 2):
    attempt = 0
    while attempt < max_attempts:
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            with open(path, "wb") as f:
                f.write(b)
            return path
        except PermissionError:
            attempt += 1
    alt_dir = os.path.join(tempfile.gettempdir(), "lipstick_index_fallback")
    os.makedirs(alt_dir, exist_ok=True)
    alt_path = os.path.join(alt_dir, os.path.basename(path))
    with open(alt_path, "wb") as f:
        f.write(b)
    print(f"WARNING: Could not write to {path} (possibly locked). Saved instead to: {alt_path}")
    return alt_path

# ----------------------------
# ABS: find & fetch Table 1 and Table 12
# ----------------------------
def fetch_abs_latest():
    soup = _get_soup(ABS_LATEST_URL)

    # Collect all xlsx/xls links on page
    xlsx_links = []
    for a in soup.select("a[href$='.xlsx'], a[href$='.xls']"):
        text = (a.get_text() or "").strip().lower()
        href = a.get("href") or ""
        if href:
            xlsx_links.append((text, _abs_make_absolute(href)))
    if not xlsx_links:
        # fallback: broader scan
        for a in soup.find_all("a"):
            href = a.get("href") or ""
            if href and href.lower().endswith((".xlsx", ".xls")):
                text = (a.get_text() or "").strip().lower()
                xlsx_links.append((text, _abs_make_absolute(href)))

    if not xlsx_links:
        raise RuntimeError("No .xlsx links found on ABS latest release page.")

    # 1) Deterministic filename selection (ABS packs often follow this scheme)
    t1_url = None     # Table 1:  by industry group
    t12_url = None    # Table 12: state by industry subgroup (SA)
    for _, url in xlsx_links:
        fname = url.split('/')[-1].lower()
        if fname == "850101.xlsx":
            t1_url = url
        if fname == "8501012.xlsx":
            t12_url = url

    # 2) Fallback: inspect workbook content & score
    if not t1_url or not t12_url:
        best_t1, best_t12 = None, None
        best_t1_score, best_t12_score = -1, -1
        for _, url in xlsx_links:
            try:
                resp = _get(url, timeout=60)
                resp.raise_for_status()
                blob = _xlsx_text_fingerprint(resp.content)
            except Exception:
                continue
            # score for Table 1
            score_t1 = 0
            for kw in ["retail turnover", "industry group", "seasonally adjusted"]:
                if kw in blob: score_t1 += 1
            if "clothing, footwear and personal accessory retailing" in blob:
                score_t1 += 2
            # score for Table 12
            score_t12 = 0
            for kw in ["state", "industry subgroup", "seasonally adjusted"]:
                if kw in blob: score_t12 += 1
            if "pharmaceutical, cosmetic and toiletry goods retailing" in blob:
                score_t12 += 2
            if score_t1 > best_t1_score:
                best_t1_score, best_t1 = score_t1, url
            if score_t12 > best_t12_score:
                best_t12_score, best_t12 = score_t12, url
        if not t1_url:
            t1_url = best_t1
        if not t12_url:
            t12_url = best_t12

    if not t1_url or not t12_url:
        print("DEBUG: Candidate filenames on the page:",
              [u.split('/')[-1] for _, u in xlsx_links][:20])
        raise RuntimeError("Could not resolve Table 1 (850101.xlsx) and/or Table 12 (8501012.xlsx).")

    t1 = _get(t1_url, timeout=60).content
    t12 = _get(t12_url, timeout=60).content

    # Save the downloaded workbooks into data/
    _safe_write_bytes(t1, os.path.join(DATA, "table1.xlsx"))
    _safe_write_bytes(t12, os.path.join(DATA, "table12.xlsx"))

    return t1, t12

# ----------------------------
# ABS: parse the needed series from the workbooks
# ----------------------------
def parse_table1_clothing(xlsx_bytes: bytes) -> pd.DataFrame:
    """
    Extract national Clothing, footwear & personal accessory retailing
    (Seasonally Adjusted, $m) from ABS Table 1 (by industry group).
    """
    xl = pd.ExcelFile(io.BytesIO(xlsx_bytes))
    df = xl.parse("Data1", header=None)

    # Find columns that are the clothing series by label (row 0)
    clothing_cols = []
    for col in df.columns[1:]:  # skip date col 0
        label = str(df.iloc[0, col] or "").lower()
        if "clothing" in label and "footwear" in label and "personal" in label and "retail" in label:
            clothing_cols.append(col)

    if not clothing_cols:
        raise RuntimeError("Could not find a clothing series label in Table 1 (row 0).")

    # Among those, pick the one that is Seasonally Adjusted (row 2)
    sa_col = None
    for col in clothing_cols:
        ser_type = str(df.iloc[2, col] or "").strip().lower()
        if ser_type.startswith("seasonally"):
            sa_col = col
            break

    if sa_col is None:
        # Fallback: search within first 8 metadata rows
        for col in clothing_cols:
            meta_rows = " ".join(str(df.iloc[r, col] or "").lower() for r in range(0, 9))
            if "seasonally adjusted" in meta_rows:
                sa_col = col
                break

    if sa_col is None:
        raise RuntimeError("Found clothing series but not the Seasonally Adjusted variant in Table 1.")

    # Build the time series from row 10 down
    dates = pd.to_datetime(df.iloc[10:, 0], errors="coerce")
    vals  = pd.to_numeric(df.iloc[10:, sa_col], errors="coerce")
    out = pd.DataFrame({"date": dates, "value_aud_m_sa": vals}).dropna()
    # normalise to month-end
    out["date"] = out["date"].dt.to_period("M").dt.to_timestamp("M")
    out = out.drop_duplicates(subset=["date"]).sort_values("date")
    if len(out) < 3:
        raise RuntimeError("Parsed too few clothing data points from Table 1 (check workbook format).")
    return out

def parse_table12_cosmetics(xlsx_bytes: bytes) -> pd.DataFrame:
    """
    Extract national Pharmaceutical, cosmetic & toiletry goods retailing (Seasonally Adjusted, $m)
    by summing the state series from ABS Table 12 (state by industry subgroup, SA).
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")

    SUBGROUP_PATTERNS = [
        r"pharmaceutical[, ]*\s*cosmetic\s*and\s*toiletry\s*goods\s*retailing",
        r"pharmaceutical.*cosmetic.*toiletry.*retail",
    ]
    SUBGROUP_RES = [re.compile(p, re.I) for p in SUBGROUP_PATTERNS]

    def looks_like_subgroup(text: str) -> bool:
        t = (text or "").strip()
        if not t:
            return False
        return any(rx.search(t) for rx in SUBGROUP_RES)

    STATE_PAT = re.compile(
        r"\b("
        r"nsw|new south wales|vic|victoria|qld|queensland|sa|south australia|"
        r"wa|western australia|tas|tasmania|nt|northern territory|act|australian capital territory"
        r")\b", re.I
    )

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb["Data1"] if "Data1" in wb.sheetnames else wb.worksheets[0]

    max_r = ws.max_row or 0
    max_c = ws.max_column or 0
    if max_r == 0 or max_c == 0:
        raise RuntimeError("Table 12 workbook appears empty.")

    # Candidate subgroup columns
    subgroup_cols = []
    for c in range(2, max_c + 1):  # columns 2..N; col 1 is date
        label0 = ws.cell(row=1, column=c).value
        if looks_like_subgroup(str(label0 or "")):
            subgroup_cols.append(c)

    if not subgroup_cols:
        raise RuntimeError("Could not find the cosmetics subgroup label in Table 12 (row 0).")

    # Keep Seasonally Adjusted + state/territory
    state_cols = []
    for c in subgroup_cols:
        meta = " ".join(
            str(ws.cell(row=r, column=c).value or "").lower()
            for r in range(1, min(10, max_r) + 1)
        )
        if ("seasonally adjusted" in meta) and STATE_PAT.search(meta):
            state_cols.append(c)

    if not state_cols:
        # fallback: accept SA even if state not detected
        for c in subgroup_cols:
            meta = " ".join(
                str(ws.cell(row=r, column=c).value or "").lower()
                for r in range(1, min(10, max_r) + 1)
            )
            if "seasonally adjusted" in meta:
                state_cols.append(c)

    if not state_cols:
        raise RuntimeError("Found cosmetics subgroup but not Seasonally Adjusted state columns in Table 12.")

    # Dates from column 1 (rows 11..)
    dates = []
    for r in range(11, max_r + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            dates.append(None)
        else:
            try:
                d = pd.to_datetime(v)
            except Exception:
                try:
                    d = pd.to_datetime(str(v))
                except Exception:
                    d = pd.NaT
            dates.append(d)

    out_df = pd.DataFrame({"date": dates})
    for c in state_cols:
        vals = []
        for r in range(11, max_r + 1):
            v = ws.cell(row=r, column=c).value
            vals.append(pd.to_numeric(v, errors="coerce"))
        out_df[f"state_{c}"] = vals

    out_df["date"] = pd.to_datetime(out_df["date"], errors="coerce")
    out_df = out_df.dropna(subset=["date"])

    state_cols_names = [col for col in out_df.columns if col.startswith("state_")]
    out_df["value_aud_m_sa"] = out_df[state_cols_names].sum(axis=1, skipna=True)

    out = out_df[["date", "value_aud_m_sa"]].copy()
    out["date"] = out["date"].dt.to_period("M").dt.to_timestamp("M")
    out = out.groupby("date", as_index=False)["value_aud_m_sa"].sum().sort_values("date")
    if len(out) < 3:
        raise RuntimeError("Parsed too few cosmetics data points from Table 12 (check workbook format).")
    return out

# ----------------------------
# ASX 200 End-of-Month (official page + many free fallbacks)
# ----------------------------
def fetch_asx200_eom_from_asx() -> pd.DataFrame:
    soup = _get_soup(ASX_HIST_URL)
    df_list = []
    # The page usually has <h3>YEAR</h3> followed by a table with monthly rows
    for header in soup.select("h2, h3, h4"):
        ytxt = header.get_text(strip=True)
        if re.fullmatch(r"\d{4}", ytxt):
            current_year = ytxt
            tbl = header.find_next("table")
            if not tbl:
                continue
            for tr in tbl.select("tr"):
                tds = [td.get_text(strip=True) for td in tr.select("td")]
                if len(tds) >= 3 and re.match(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)$', tds[0]):
                    month, _allords, asx200 = tds[0], tds[1], tds[2]
                    try:
                        asx200_val = float(asx200.replace(',', ''))
                    except Exception:
                        continue
                    date = eom_date_from_text(month, current_year)
                    df_list.append((date, asx200_val))
    if not df_list:
        raise RuntimeError("Could not parse ASX 200 EOM data from ASX page.")
    out = pd.DataFrame(df_list, columns=['date','asx200_eom_index']).sort_values('date')
    out = out.drop_duplicates(subset=["date"])
    return out

def fetch_asx200_eom_fallback_yahoo(max_retries: int = 5, base_sleep: float = 2.0) -> pd.DataFrame:
    """Yahoo fallback with heavy-duty retries, multi-host rotation, and alternate HTML-scrape path."""
    import time

    # ------------- CSV endpoint (preferred) -------------
    def _try_yahoo_csv(host: str) -> pd.DataFrame | None:
        import time as _t
        period1 = 946684800  # 2000-01-01
        period2 = int(_t.time())
        url = (f"https://{host}/v7/finance/download/%5EAXJO"
               f"?period1={period1}&period2={period2}&interval=1mo&events=history&includeAdjustedClose=true")
        try:
            r = _get(url, timeout=30)
            if r.status_code in (429, 503):
                return None
            r.raise_for_status()
            df = pd.read_csv(io.StringIO(r.text))
            if "Date" not in df or "Close" not in df:
                return None
            df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
            df = df.dropna(subset=["Date", "Close"])
            df = df.rename(columns={"Date":"date", "Close":"asx200_eom_index"})
            df["date"] = df["date"].dt.to_period("M").dt.to_timestamp("M")
            df = df.drop_duplicates(subset=["date"]).sort_values("date")
            return df[["date","asx200_eom_index"]]
        except Exception:
            return None

    # ------------- HTML history page (secondary) -------------
    def _try_yahoo_history_html() -> pd.DataFrame | None:
        import re as _re
        try:
            r = _get(YF_HTML_URL, timeout=30)
            if r.status_code in (429, 503):
                return None
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "html.parser")
            rows = []
            for tr in soup.select("table tbody tr"):
                tds = [td.get_text(strip=True).replace("\u2212", "-") for td in tr.select("td")]
                if len(tds) >= 6 and _re.search(r"\w{3}\s+\d{1,2},\s+\d{4}", tds[0]):
                    try:
                        d = pd.to_datetime(tds[0])
                        close = float(tds[4].replace(',', ''))
                    except Exception:
                        continue
                    rows.append((d.to_period("M").to_timestamp("M"), close))
            if not rows:
                return None
            df = pd.DataFrame(rows, columns=["date","asx200_eom_index"]).drop_duplicates("date").sort_values("date")
            return df
        except Exception:
            return None

    for attempt in range(max_retries):
        random.shuffle(YF_CSV_HOSTS)
        for h in YF_CSV_HOSTS:
            df = _try_yahoo_csv(h)
            if df is not None:
                return df
        df = _try_yahoo_history_html()
        if df is not None:
            return df
        sleep_s = base_sleep * (2 ** attempt) + random.uniform(0, 0.6)
        time.sleep(sleep_s)

    raise RuntimeError("Yahoo CSV/HTML paths unavailable.")

def fetch_asx200_eom_fallback_stooq() -> pd.DataFrame:
    """Stooq monthly CSV fallbacks."""
    stooq_urls = [
        "https://stooq.com/q/d/l/?s=%5Eaxjo&i=m",
        "https://stooq.com/q/d/l/?s=xjo&i=m",
    ]
    for url in stooq_urls:
        try:
            r = _get(url, timeout=30)
            r.raise_for_status()
            sdf = pd.read_csv(io.StringIO(r.text))
            if "Date" in sdf.columns and "Close" in sdf.columns:
                sdf["Date"] = pd.to_datetime(sdf["Date"], errors="coerce")
                sdf = sdf.dropna(subset=["Date", "Close"])
                sdf = sdf.rename(columns={"Date":"date", "Close":"asx200_eom_index"})
                sdf["date"] = sdf["date"].dt.to_period("M").dt.to_timestamp("M")
                sdf = sdf.drop_duplicates(subset=["date"]).sort_values("date")
                if not sdf.empty:
                    return sdf[["date","asx200_eom_index"]]
        except Exception:
            continue
    raise RuntimeError("Stooq monthly CSV unavailable.")

def fetch_asx200_eom_fallback_marketwatch() -> pd.DataFrame:
    """
    MarketWatch CSV: pull DAILY history via 'downloaddatapartial' for a long range, then
    compute month-end closes. Endpoint documented by usage; parameters:
      csvdownload=true, frequency=p1d, daterange=y50 (50 years)
    """
    params = {
        "countrycode": "au",
        "csvdownload": "true",
        "downloadpartial": "false",
        "frequency": "p1d",
        "newdates": "false",
        "daterange": "y50",
    }
    try:
        r = _get(MW_CSV_BASE, params=params, timeout=45)
        if r.status_code in (429, 503):
            raise RuntimeError("MarketWatch throttled")
        r.raise_for_status()
        # Parse CSV safely (handles quoted numbers with commas)
        reader = csv.DictReader(io.StringIO(r.text))
        rows = []
        for row in reader:
            try:
                d = pd.to_datetime(row["Date"], errors="coerce")
                c = float(row["Close"].replace(",", ""))
            except Exception:
                continue
            if pd.notna(d):
                rows.append((d, c))
        if not rows:
            raise RuntimeError("MarketWatch CSV parsed 0 rows")
        df = pd.DataFrame(rows, columns=["date", "close"]).sort_values("date")
        # Resample to month-end using last available trading day close
        df["date"] = df["date"].dt.to_period("M").dt.to_timestamp("M")
        df = df.groupby("date", as_index=False)["close"].last()
        df = df.rename(columns={"close": "asx200_eom_index"})
        if df.empty:
            raise RuntimeError("MarketWatch produced empty monthly frame")
        return df[["date", "asx200_eom_index"]]
    except Exception as e:
        raise RuntimeError(f"MarketWatch CSV unavailable ({e}).")

def fetch_asx200_eom_fallback_investing() -> pd.DataFrame:
    """
    Investing.com Monthly HTML table: scrape the 'Historical Data' page for the AUS 200 index.
    We read the table (date, price) where date strings are monthly. Returns EOM series.
    """
    try:
        r = _get(INV_HTML_URL, timeout=45, headers={
            "User-Agent": _ua(),
            "Accept-Language": "en-US,en;q=0.9",
        })
        if r.status_code in (429, 503):
            raise RuntimeError("Investing.com throttled")
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        # Find the main historical-data table
        # The page often renders a table with rows like: "Aug 2025 8,662.00 ..."
        tbl = None
        for t in soup.select("table"):
            head = " ".join(th.get_text(strip=True).lower() for th in t.select("thead th"))
            if ("date" in head) and ("price" in head or "close" in head):
                tbl = t
                break
        if tbl is None:
            # Fallback: any table with at least 2 columns where first looks like a date
            for t in soup.select("table"):
                tbl = t
                break
        rows = []
        for tr in (tbl.select("tbody tr") if tbl else []):
            tds = [td.get_text(strip=True) for td in tr.select("td")]
            if len(tds) < 2:
                continue
            # Accept a wide range of date formats (e.g., "Aug 2025", "Jul 31, 2025")
            d = pd.to_datetime(tds[0], errors="coerce")
            if pd.isna(d):
                continue
            price_txt = re.sub(r"[,\s]", "", tds[1])
            try:
                px = float(price_txt)
            except Exception:
                # Some tables use a different column order (e.g., Close is 2nd or 5th)
                found = None
                for cell in tds[1:6]:
                    cell_clean = re.sub(r"[,\s]", "", cell)
                    try:
                        found = float(cell_clean)
                        break
                    except Exception:
                        continue
                if found is None:
                    continue
                px = found
            rows.append((d.to_period("M").to_timestamp("M"), px))
        if not rows:
            raise RuntimeError("Could not parse monthly rows from Investing.com")
        df = pd.DataFrame(rows, columns=["date","asx200_eom_index"]).drop_duplicates("date").sort_values("date")
        return df
    except Exception as e:
        raise RuntimeError(f"Investing.com HTML unavailable ({e}).")

def fetch_asx200_eom() -> pd.DataFrame:
    """
    Try, in order:
      ASX official → Yahoo (robust) → Stooq → MarketWatch → Investing.com
    If all fail and a local cache exists (data/asx200_eom.csv), return that instead.
    If there is no cache, create a fill-in template and raise a friendly error.
    """
    # 1) ASX official
    try:
        return fetch_asx200_eom_from_asx()
    except Exception as e:
        print(f"WARNING: ASX official page parse failed ({e}). Trying Yahoo…")

    # 2) Yahoo with backoff (CSV/HTML)
    try:
        return fetch_asx200_eom_fallback_yahoo()
    except Exception as e:
        print(f"WARNING: Yahoo fallback failed ({e}). Trying Stooq…")

    # 3) Stooq
    try:
        return fetch_asx200_eom_fallback_stooq()
    except Exception as e:
        print(f"WARNING: Stooq fallback failed ({e}). Trying MarketWatch…")

    # 4) MarketWatch (daily CSV → monthly)
    try:
        return fetch_asx200_eom_fallback_marketwatch()
    except Exception as e:
        print(f"WARNING: MarketWatch fallback failed ({e}). Trying Investing.com…")

    # 5) Investing.com (monthly HTML)
    try:
        return fetch_asx200_eom_fallback_investing()
    except Exception as e:
        print(f"WARNING: Investing.com fallback failed ({e}). Checking local cache…")

    # 6) Local cache fallback
    cache_path = os.path.join(DATA, "asx200_eom.csv")
    if os.path.exists(cache_path):
        try:
            c = pd.read_csv(cache_path)
            c["date"] = pd.to_datetime(c["date"])
            c = c[["date", "asx200_eom_index"]].dropna()
            c["date"] = c["date"].dt.to_period("M").dt.to_timestamp("M")
            c = c.drop_duplicates(subset=["date"]).sort_values("date")
            if not c.empty:
                print("Using cached data/asx200_eom.csv as a fallback.")
                return c
        except Exception:
            pass

    # 7) Create a template to fill in manually once, so future runs work offline
    tmpl = pd.DataFrame({
        "date": pd.date_range("2019-01-31", periods=24, freq="ME"),
        "asx200_eom_index": [np.nan]*24
    })
    _safe_to_csv(tmpl, cache_path)
    raise RuntimeError(
        "ASX200 sources are unavailable right now. A fill-in template was created at "
        f"{cache_path}. Paste end-of-month S&P/ASX 200 levels (price index) into the file "
        "and re-run. The script will use your cached file thereafter."
    )

# ----------------------------
# Core computations
# ----------------------------
def compute_mom_pct(df: pd.DataFrame, valcol: str, newcol: str) -> pd.DataFrame:
    df = df.copy().sort_values("date")
    df[newcol] = df[valcol].pct_change() * 100.0
    return df[["date", newcol]]

def build_index(cosmetics_mom: pd.DataFrame, clothing_mom: pd.DataFrame) -> pd.DataFrame:
    df = pd.merge(
        cosmetics_mom[["date","cosmetics_mom"]],
        clothing_mom[["date","clothing_mom"]],
        on="date", how="inner"
    )
    df["lipstick_index"] = WEIGHT_COSMETICS*df["cosmetics_mom"] + WEIGHT_CLOTHING*df["clothing_mom"]
    return df

def compute_asx_returns(asx_df: pd.DataFrame) -> pd.DataFrame:
    asx_df = asx_df.copy().sort_values("date")
    asx_df["asx200_ret_mom"] = asx_df["asx200_eom_index"].pct_change() * 100.0
    return asx_df[["date","asx200_ret_mom"]]

def find_divergences(li_df: pd.DataFrame, asx_ret: pd.DataFrame):
    merged = pd.merge(li_df, asx_ret, on="date", how="inner")
    cond = (merged["lipstick_index"] > 0.0) & (merged["asx200_ret_mom"] < 0.0)
    out = merged.loc[cond, ["date","cosmetics_mom","clothing_mom","lipstick_index","asx200_ret_mom"]].copy()
    out = out.sort_values("date").reset_index(drop=True)
    return merged, out

def save_charts(merged_df: pd.DataFrame):
    # Time series
    plt.figure(figsize=(10,5))
    plt.plot(merged_df['date'], merged_df['lipstick_index'], label='Lipstick Index (MoM, %)')
    plt.plot(merged_df['date'], merged_df['asx200_ret_mom'], label='ASX200 Return (MoM, %)')
    plt.legend()
    plt.title("Lipstick Index vs ASX200 Monthly Returns")
    plt.xlabel("Date")
    plt.ylabel("%")
    plt.tight_layout()
    ts_path = os.path.join(OUT, "lipstick_vs_asx_timeseries.png")
    plt.savefig(ts_path, dpi=160)
    plt.close()

    # Scatter
    plt.figure(figsize=(6,6))
    plt.scatter(merged_df['asx200_ret_mom'], merged_df['lipstick_index'])
    plt.axhline(0); plt.axvline(0)
    plt.xlabel("ASX200 MoM return (%)")
    plt.ylabel("Lipstick Index MoM (%)")
    plt.title("Lipstick vs ASX200 (monthly)")
    plt.tight_layout()
    sc_path = os.path.join(OUT, "lipstick_scatter.png")
    plt.savefig(sc_path, dpi=160)
    plt.close()

    return ts_path, sc_path

# ----------------------------
# Main
# ----------------------------
def main():
    print(">>")
    print("Fetching ABS latest Table 1 & 12...")
    t1_bytes, t12_bytes = fetch_abs_latest()

    print("Parsing clothing (Table 1)...")
    clothing = parse_table1_clothing(t1_bytes)
    print(f"Clothing rows: {len(clothing)}")

    print("Parsing cosmetics (Table 12)...")
    cosmetics = parse_table12_cosmetics(t12_bytes)
    print(f"Cosmetics rows: {len(cosmetics)}")

    # Save cleaned series (for reuse) into data/
    _safe_to_csv(clothing, os.path.join(DATA, "clothing_sa.csv"))
    _safe_to_csv(cosmetics, os.path.join(DATA, "cosmetics_sa.csv"))

    print("Fetching ASX 200 EOM list...")
    asx = fetch_asx200_eom()
    _safe_to_csv(asx, os.path.join(DATA, "asx200_eom.csv"))

    # Build index and divergences
    cos_m = compute_mom_pct(cosmetics, "value_aud_m_sa", "cosmetics_mom")
    clo_m = compute_mom_pct(clothing,  "value_aud_m_sa", "clothing_mom")
    li_df = build_index(cos_m, clo_m)

    asx_ret = compute_asx_returns(asx)
    merged, divergences = find_divergences(li_df, asx_ret)

    # Save outputs
    out_csv = os.path.join(DATA, "lipstick_divergences.csv")
    _safe_to_csv(divergences, out_csv)

    ts_path, sc_path = save_charts(merged)

    # Pretty print
    if divergences.empty:
        print("No divergence months found (Lipstick > 0 & ASX200 < 0) for the overlapping period.")
    else:
        print("Divergence months (Lipstick up, ASX200 down):")
        print(divergences.to_string(
            index=False,
            formatters={
                'cosmetics_mom': '{:.2f}'.format,
                'clothing_mom': '{:.2f}'.format,
                'lipstick_index': '{:.2f}'.format,
                'asx200_ret_mom': '{:.2f}'.format,
            }
        ))

    print(f"\nSaved CSVs in: {DATA}")
    print(f"Charts in: {OUT}")
    print("Done.")

if __name__ == "__main__":
    main()
